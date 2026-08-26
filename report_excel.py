#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
report_excel.py — 검측 결과를 엑셀 조서로
========================================================================
입력 이미지 한 장과 설계값(사양 프로파일)을 넣어 얻은 결과를 하나의
.xlsx 로 낸다. 시트 구성은 읽는 순서를 따른다.

  1. 요약            무엇을 어떤 설계값으로 쟀고 결론이 무엇인가
  2. 설계값          삼각측량식에 들어간 값과 그 출처 등급
  3. 세그멘테이션      색깔별로 무엇을 무엇으로 구분했는가
  4. 검측결과         구분별 수직·수평·평활도 판정
  5. 평활도 상세      KCS 직선자 길이별 처짐과 허용치
  6. 유의사항         판정을 그대로 믿으면 안 되는 항목

왜 색 범례를 따로 두는가
  세그멘테이션 오버레이 이미지만 보면 파란 면이 벽인지 거푸집인지,
  주황 선이 동바리인지 기둥인지 알 수 없다. 그런데 그 구분에 따라
  적용되는 KCS 허용오차가 달라진다(동바리와 철근은 기준이 다르다).
  그래서 색 → 클래스 → 적용 기준을 한 시트에 묶어 둔다. 기하 전용
  백엔드는 이 구분을 못 하므로 그 사실도 같은 시트에 적는다.
========================================================================
"""
import numpy as np
import importlib.util as _ilu, os as _os
import tempfile as _tempfile
import atexit as _atexit


def _load(name):
    spec = _ilu.spec_from_file_location(
        name, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                            f"{name}.py"))
    m = _ilu.module_from_spec(spec); spec.loader.exec_module(m)
    return m


REPORT = _load("report")
CALIB = _load("calibration")
EQ3 = _load("eq3_orientation")

CLASS_KO = REPORT.CLASS_KO
CLASS_COLOR = REPORT.CLASS_COLOR
KIND_KO = REPORT.KIND_KO

# 부재별로 실제 적용되는 KCS 조항. 세그멘테이션이 무엇으로 봤는지에 따라
# 이 기준이 갈리므로 범례 시트에 함께 낸다.
KCS_NOTE = {
    "wall":            "KCS 14 20 10 — 위치 ±20mm, 단면 −5/+20mm",
    "formwork_wall":   "KCS 14 20 12 — 공사시방서별(일반 ±20mm)",
    "formwork_column": "KCS 14 20 12 — 공사시방서별(일반 ±20mm)",
    "masonry":         "KCS 41 34 02 — 공사시방서별(일반 ±10mm/층)",
    "plaster_wall":    "KCS 41 46 00 — 공사시방서별",
    "column":          "KCS 14 20 10 / 41 30 02 — ±20mm, 권장 ±h/1000",
    "floor":           "KCS 14 20 10 — 위치 ±20mm, 두께 −5/+20mm",
    "slab":            "KCS 14 20 10 — 위치 ±20mm",
    "ceiling":         "KCS 41 52 00 — 일반 3m당 ±3mm",
    "shoring":         "가설 동바리 — 연직 정밀도, 공사시방서별",
    "rebar":           "철근 조립 — 연직 정밀도, 공사시방서별",
    "background":      "검측 대상 아님",
}

_HDR_BG = "1F2937"
_HDR_FG = "FFFFFF"
_SEC_BG = "E5E7EB"


def _rgb_hex(rgb):
    return "%02X%02X%02X" % tuple(int(c) for c in rgb)


def _style(ws, openpyxl, widths=None, header_row=1):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="D1D5DB")
    for c in ws[header_row]:
        if c.value is None:
            continue
        c.font = Font(bold=True, color=_HDR_FG, size=10)
        c.fill = PatternFill("solid", fgColor=_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    for row in ws.iter_rows(min_row=header_row):
        for c in row:
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if c.row > header_row:
                c.alignment = Alignment(vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _verdict_fill(openpyxl, value):
    from openpyxl.styles import PatternFill, Font
    col = {"합격": ("DCFCE7", "166534"), "기준초과": ("FEE2E2", "991B1B"),
           "판정보류(분해능)": ("FEF9C3", "854D0E"),
           "판정보류(노출길이)": ("FEF9C3", "854D0E"),
           "판정보류(단면 미분해)": ("FEF9C3", "854D0E"),
           "측정불가": ("F1F5F9", "475569"), "해당없음": ("F8FAFC", "94A3B8")}
    return col.get(value)


# =====================================================================
# 시트별 작성
# =====================================================================
# 그림을 줄여 담아 둔 임시 파일. openpyxl 이 wb.save() 시점에 읽으므로
# 그때까지 살아 있어야 하고, 프로세스가 끝나면 지운다.
_EMBED_TMP = []


@_atexit.register
def _cleanup_embed_tmp():
    for _f in _EMBED_TMP:
        try:
            _os.unlink(_f)
        except OSError:
            pass


def _embed_image(ws, path, row, max_px=1400, col="A", caption=None,
                 openpyxl=None):
    """
    그림을 시트에 붙인다.

    원본을 그대로 넣지 않고 **줄여서** 넣는다. 2448×2048 세그멘테이션
    이미지는 3.7MB 라, 시트 셋에 원본을 넣으면 조서 하나가 7.5MB 가 된다.
    엑셀은 파일에 담긴 화소를 통째로 안고 가며 표시 크기만 줄여도 용량은
    그대로다. 현장에서 열어 보는 조서로는 무겁다.

    max_px 는 긴 변 기준이다. 1400px 이면 화면에서 확대해도 격자점이
    구분되고, 용량은 원본의 1/5~1/10 이 된다.
    """
    if not path or not _os.path.exists(path):
        ws.cell(row=row, column=1,
                value=f"[그림 없음] {path or '경로 미지정'}")
        return row + 1
    try:
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
    except ImportError:
        ws.cell(row=row, column=1, value="[그림 생략] Pillow/openpyxl 필요")
        return row + 1
    try:
        if caption:
            c = ws.cell(row=row, column=1, value=caption)
            if openpyxl is not None:
                c.font = openpyxl.styles.Font(bold=True, size=11)
            row += 1
        with PILImage.open(path) as im:
            im = im.convert("RGB")
            w, h = im.size
            sc = min(1.0, float(max_px) / max(w, h))
            if sc < 1.0:
                im = im.resize((max(1, int(w * sc)), max(1, int(h * sc))),
                               PILImage.LANCZOS)
            tmp = _tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            tmp.close()
            im.save(tmp.name, format="PNG", optimize=True)
            _EMBED_TMP.append(tmp.name)
            iw, ih = im.size
        xi = XLImage(tmp.name)
        xi.width, xi.height = iw, ih
        ws.add_image(xi, f"{col}{row}")
        # 그림이 덮는 만큼 행을 비워 둔다 (기본 행 높이 20px 가정)
        return row + int(ih / 19.0) + 3
    except Exception as e:
        ws.cell(row=row, column=1, value=f"[그림 삽입 실패] {e}")
        return row + 1


def _sheet_summary(wb, openpyxl, record, meta, result):
    ws = wb.active
    ws.title = "1.요약"
    s = record["summary"]
    rows = [("항목", "값")]
    for k, v in meta.items():
        rows.append((k, v))
    rows += [
        ("", ""),
        ("설계값 프로파일", f"{CALIB.ACTIVE_PROFILE} — "
                      f"{CALIB.SPEC_PROFILES[CALIB.ACTIVE_PROFILE]['label']}"),
        ("세그멘테이션 백엔드", s.get("세그멘테이션")),
        ("", ""),
        ("검출 영역 수", s.get("영역수")),
        ("검측 완료", s.get("검측")),
        ("기각", s.get("기각")),
        ("라벨 교정", s.get("라벨교정")),
        ("영역 재분할", s.get("영역재분할")),
        ("", ""),
        ("자세(수직·수평) 기준초과", s.get("자세_기준초과")),
        ("평활도 기준초과", s.get("평활도_기준초과")),
        ("평활도 판정보류", s.get("평활도_판정보류")),
    ]
    g = s.get("중력벡터")
    if g:
        rows.append(("조사기 좌표계 중력벡터",
                     "(" + ", ".join(f"{x:.4f}" for x in g) + ")"))
    for r in rows:
        ws.append(list(r))
    _style(ws, openpyxl, widths=[26, 62])
    from openpyxl.styles import Font
    for row in ws.iter_rows(min_row=2, max_col=1):
        row[0].font = Font(bold=True, size=10)
    return ws


def _sheet_design(wb, openpyxl):
    """삼각측량식에 실제로 들어간 값과 출처 등급."""
    ws = wb.create_sheet("2.설계값")
    ws.append(["항목", "기호", "값", "출처", "근거 / 필요한 조치"])
    prov = CALIB.provenance()
    ang = CALIB.make_line_angles()
    a0 = np.degrees(ang["V0"]["angle_rad"])
    a1 = np.degrees(ang[f"V{CALIB.N_VERTICAL-1}"]["angle_rad"])
    rows = [
        ("카메라 초점거리", "f", f"{CALIB.F_PX:.1f} px", "f_px"),
        ("주점", "c_x, c_y", f"{CALIB.CX_PX:.1f}, {CALIB.CY_PX:.1f} px", "cx_px"),
        ("기선", "b", f"{CALIB.BASELINE_M*1000:.0f} mm", "b_m"),
        ("DOE 발산각", "—", f"{CALIB.FOV_DEG:.2f}°", "fov_deg"),
        ("격자선 수", "—",
         f"수직 {CALIB.N_VERTICAL} + 수평 {CALIB.N_HORIZONTAL} "
         f"({CALIB.N_VERTICAL*CALIB.N_HORIZONTAL} 교점)", "n_lines"),
        ("센서", "—",
         f"{CALIB.IMAGE_W}×{CALIB.IMAGE_H} @{CALIB.PIXEL_PITCH_UM}µm "
         f"{CALIB.SENSOR_COLOR}"
         + (f" + {CALIB.OPTICAL_FILTER_NM}nm 대역통과"
            if CALIB.OPTICAL_FILTER_NM else ""), "sensor"),
        ("레이저 수렴각", "δ", f"{CALIB.LASER_TILT_DEG}°", "tilt"),
        ("V선 발사각", "α_i",
         f"{CALIB.DOE_ANGLE_MODEL} {CALIB.N_VERTICAL}분할 "
         f"({a0:+.2f}° … {a1:+.2f}°)", "alpha_i"),
        ("H선 발사각", "β_j",
         f"{CALIB.DOE_ANGLE_MODEL} {CALIB.N_HORIZONTAL}분할", "beta_j"),
        ("카메라–레이저 자세", "R, t", "R=I, t=(b,0,0)", "R_t"),
        ("IMU–카메라 자세", "R_ic", "단위행렬", "R_ic"),
        ("가속도계 bias", "b_a", "미구현", "b_a"),
        ("선검출 픽셀오차", "σ_u", f"{CALIB.SIGMA_U_PX} px", "sigma_u"),
    ]
    from openpyxl.styles import PatternFill, Font
    grade_fill = {"spec": ("DCFCE7", "166534"),
                  "design": ("DBEAFE", "1E40AF"),
                  "assumed": ("FEF9C3", "854D0E")}
    for name, sym, val, key in rows:
        grade, note = prov[key]
        ws.append([name, sym, val, grade, note])
        c = ws.cell(row=ws.max_row, column=4)
        bg, fg = grade_fill[grade]
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(bold=True, color=fg, size=10)

    ws.append([])
    ws.append(["파생값", "", "", "", ""])
    z = 1.2
    ws.append(["깊이 잡음 σ_Z @1.2m", "σ_Z", f"{CALIB.sigma_z_mm(z):.3f} mm",
               "", "σ_u·Z²/(f·b)"])
    ws.append(["격자 피치 @1.2m", "—",
               f"{CALIB.projection_mm_at(z)/(CALIB.N_VERTICAL-1):.1f} mm", "",
               "이 값보다 가는 부재는 표본이 한두 점뿐"])
    near, far = CALIB.depth_of_field()
    ws.append(["고정초점 심도", "—", f"{near:.2f} ~ {far:.2f} m", "",
               f"작업거리 {CALIB.WORK_Z_MIN_M}~{CALIB.WORK_Z_MAX_M}m 를 덮어야 한다"])
    _style(ws, openpyxl, widths=[22, 10, 34, 10, 52])
    return ws


def _sheet_detection(wb, openpyxl, det, e2e=None, tri=None,
                     overlay_image=None):
    """
    선검출 정확도 — 화소가 맞아야 3D 가 맞는다.

    삼각측량은 Z = f·b/(f·tanα − (u−c_x)) 라 분모가 화소 차이다. 화소가
    1px 흔들리면 깊이가 Z²/(f·b) 만큼 흔들린다. 뒤쪽 검측식이 아무리
    정확해도 여기서 끝난다. 그래서 조서에 넣는다.
    """
    ws = wb.create_sheet("3.선검출(1단계)")
    from openpyxl.styles import PatternFill, Font
    if not det:
        ws.append(["선검출 정확도", "평가하지 않음"])
        ws.append(["", "정답 화소가 있는 내보내기(raycast) 입력에서만 잴 수 있다. "
                       "실촬영에는 대조할 정답이 없다."])
        _style(ws, openpyxl, widths=[22, 100])
        return ws
    if det.get("error"):
        ws.append(["선검출 정확도", det["error"]])
        _style(ws, openpyxl, widths=[22, 100])
        return ws

    mm = det.get("mm_per_px_depth")
    ws.append(["항목", "값", "깊이 환산", "비고"])
    rows = []
    if tri:
        rows += [
            ("[0단계] 삼각측량 자체", "", "",
             "정답 화소를 3D 로 풀어 내보내기의 xyz_world 와 맞댄 것. "
             "검출이 아니라 계산(f·주점·기선·자세·eq1)이 맞는지 본다"),
            ("3D 위치 오차", f"{tri['n_points']:,}점",
             f"{tri['dist_med_mm']:.4f} mm",
             f"최대 {tri['dist_max_mm']:.3f} mm — 계산 체인에 문제 없음"),
            ("", "", "", ""),
        ]
    rows += [
        ("대조 이미지", det["image"], "",
         f"{det['image_size'][0]}×{det['image_size'][1]} "
         f"(센서 {det['f_px_sensor']:.0f}px 기준 배율 {det['scale_to_sensor']})"),
        ("대표 측정거리", f"{det['z_ref_m']} m", "",
         f"화소 1px 이 깊이 {mm} mm 에 해당"),
        ("", "", "", ""),
        ("V선 검출", f"{det['n_v_matched']} / {det['n_v_total']}", "",
         ("미검출: " + ", ".join(det["missed_lines"]))
         if det["missed_lines"] else "전부 검출"),
        ("V선 번호 일치", f"{det['id_ok']['V'][0]} / {det['id_ok']['V'][1]}", "",
         "선 번호가 발사각 α 를 정한다 — 틀리면 깊이가 통째로 어긋난다"),
        ("H선 번호 일치", f"{det['id_ok']['H'][0]} / {det['id_ok']['H'][1]}", "",
         "가로선 번호. 아래 계열별 표에서 깊이 기여를 따로 본다"),
        ("", "", "", ""),
    ]

    # ── 계열별 (세로 V / 가로 H) ──
    # 검출이 잘 되는가와 깊이를 줄 수 있는가는 별개 문제다. 가로선은
    # 완벽하게 검출되면서도 깊이는 하나도 못 줄 수 있고, 그 원인은
    # 검출이 아니라 기하(eq7 의 이득 g)에 있다. 둘을 한 표에 나란히
    # 놓아야 "가로축을 왜 못 쓰는가" 가 조서에서 읽힌다.
    fam = det.get("families") or {}
    if fam:
        rows += [("[계열별] 세로선 vs 가로선", "", "",
                  "검출 정확도와 깊이 기여를 나눠 본다")]
        for pre, d in fam.items():
            gm = d["깊이이득_중앙"]
            sw = d["정답선_변동폭_px"]
            usable = d["깊이가능"]
            rows.append((
                f"{d['이름']} 검출",
                f"{d['검출']} / {d['선 수']}",
                "",
                (f"{d['측정축']} 좌표 기준. 선내잡음 {d['선내잡음_px']} px "
                 f"+ 선간편차 {d['선간편차_px']} px → 통합 {d['통합오차_px']} px")))
            rows.append((
                f"{d['이름']} 깊이 신호",
                (f"{sw:.4f} px" if sw is not None else "-"),
                "",
                ("정답선이 측정축으로 실제 움직이는 폭. 이 폭이 검출 잡음보다 "
                 "작으면 깊이를 읽어낼 것이 없다")))
            rows.append((
                f"{d['이름']} 깊이 사용",
                f"{usable} / {d['선 수']}",
                "",
                (f"레이저 평면 이득 g ≈ {gm} (기준 g ≤ "
                 f"{det.get('max_depth_gain', 3.0):g})"
                 if gm is not None else
                 "g = ∞ — 레이저 평면이 기선(X축)을 품어 시차가 선과 나란하다. "
                 "검출을 아무리 잘해도 깊이가 안 풀린다. 격자를 광축 둘레로 "
                 "굴려야(roll) 풀린다 — calibration 의 diagonal 프로파일 참조")))
        rows += [("", "", "", "")]

    rows += [
        ("[화소 오차] 무엇을 뺐나", "검출 u − 정답 u", "",
         "같은 행(v)에서 선검출이 찾은 u 와 raycast 정답 u 의 차이. "
         "깊이가 아니라 화면 좌표 차이다"),
        ("계통 편차 (중앙값)", f"{det['err_bias_px']:+.3f} px",
         f"{det['depth_bias_mm']:.1f} mm",
         "모든 점이 한쪽으로 치우친 몫. 좌표 규약·주점·기선이 어긋나면 "
         "생기며 소프트웨어로 제거할 수 있다"),
        ("무작위 오차 (σ)", f"{det['err_noise_px']:.3f} px",
         f"{det['depth_noise_mm']:.1f} mm",
         f"치우침을 뺀 나머지의 표준편차 = 점마다 제각각 흔들리는 몫. "
         f"이것이 실제 검출 정밀도 σ_u 다 (설계 가정 "
         f"{det['sigma_u_design_px']} px). 깊이 환산은 dZ = Z²/(f·b)·du"),
        ("전체 RMS", f"{det['err_rms_px']:.3f} px",
         f"{det['depth_err_mm']:.1f} mm", "계통 + 무작위"),
        ("95 백분위", f"{det['err_p95_px']:.3f} px", "", ""),
    ]
    if det.get("dz_noise_mm") is not None:
        rows += [
            ("", "", "", ""),
            ("[깊이 오차] 무엇을 뺐나", "검출 Z − 정답 Z", "",
             "위의 두 화소를 각각 삼각측량해 나온 깊이의 차이. 환산이 "
             "아니라 실제로 두 번 풀어 뺀 값이다"),
            ("깊이 치우침", "", f"{det['dz_med_mm']:+.2f} mm",
             "선별 중앙값의 중앙값"),
            ("깊이 흔들림 (σ)", "", f"{det['dz_noise_mm']:.2f} mm",
             "한 격자점의 깊이가 이만큼 흔들린다는 뜻. 면적합은 이를 "
             "수만 점 평균해 지우지만, 평활도는 그대로 받는다"),
            ("깊이 95 백분위", "", f"{det['dz_p95_mm']:.2f} mm", ""),
        ]
    if det.get("quantization_px"):
        rows.append(("렌더 양자화 한계", f"{det['quantization_px']:.3f} px",
                     f"{det['quantization_mm']:.1f} mm",
                     "레이저선이 안티에일리어싱 없이 이진(0/255)으로 그려져 "
                     "있어 위치가 0.5px 격자에 갇힌다. 어떤 서브픽셀 알고리즘도 "
                     "이 아래로는 못 내려간다 — 입력을 고쳐야 하는 몫이다"))
    for r in rows:
        ws.append(list(r))
    _style(ws, openpyxl, widths=[20, 16, 13, 62])

    # 설계 가정과의 비교를 색으로
    sig = det["sigma_u_design_px"]
    c = ws.cell(row=10, column=2)
    bad = det["err_noise_px"] > sig * 1.5
    c.fill = PatternFill("solid", fgColor="FEE2E2" if bad else "DCFCE7")
    c.font = Font(bold=True, color="991B1B" if bad else "166534", size=10)

    if e2e and e2e.get("rows"):
        ws.append([])
        ws.append(["검출 화소로 검측까지 돌린 결과", "", "", ""])
        h2 = ws.max_row + 1
        ws.append(["부재", "정답 화소 결과", "검출 화소 결과", "차이"])
        for r in e2e["rows"]:
            if r.get("dtheta_deg") is None:
                ws.append([r["class"], f"{r['theta_gt']}°",
                           "측정 실패", "—"])
                continue
            ws.append([r["class"], f"각도 {r['theta_gt']}°",
                       f"각도 {r['theta_det']}°", f"{r['dtheta_deg']}°"])
            if r.get("dgap_mm") is not None:
                ws.append(["", f"자처짐 {r['gap_gt_mm']} mm",
                           f"자처짐 {r['gap_det_mm']} mm", f"{r['dgap_mm']} mm"])
        _style(ws, openpyxl, widths=[20, 16, 13, 62], header_row=h2)
        ws.append([])
        ws.append(["읽는 법",
                   "각도는 면적합이 수만 점을 평균하므로 화소 오차가 거의 "
                   "옮겨 붙지 않는다. 평활도는 점별 오차가 그대로 표면 "
                   "요철로 보이므로 화소 정밀도가 곧 결과다.", "", ""])

    ws.append([])
    ws.append(["선별 상세", "", "", ""])
    hdr = ws.max_row + 1
    ws.append(["선", "방향", "거리(m)", "정답 위치(px)", "검출 대응",
               "정답 점수", "검출 점수", "계통(px)", "σ(px)", "RMS(px)",
               "최대(px)", "비고"])
    for r in det["rows"]:
        ws.append([
            r["lid"], "수직" if r["fixed"] == "alpha" else "수평",
            r["z_m"], r.get("gt_pos"), r.get("matched"),
            r["n_gt"], r.get("n_det"),
            (round(r["err_med"], 3) if r.get("err_med") is not None else None),
            (round(r["err_noise"], 3) if r.get("err_noise") is not None else None),
            (round(r["err_rms"], 3) if r.get("err_rms") is not None else None),
            (round(r["err_max"], 2) if r.get("err_max") is not None else None),
            r.get("note") or ("번호 불일치" if not r.get("id_ok") else "")])
        if r.get("err_rms") is None:
            for col in range(1, 13):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill(
                    "solid", fgColor="FEE2E2")
    _style(ws, openpyxl,
           widths=[6, 6, 9, 13, 11, 10, 10, 10, 9, 10, 10, 40],
           header_row=hdr)
    _embed_image(ws, overlay_image, ws.max_row + 3, max_px=1600,
                 openpyxl=openpyxl,
                 caption="[그림] 선검출 대조 — 배경은 레이캐스트 렌더, "
                         "자홍색이 검출한 선, 청록색이 정답 화소")
    return ws


def _sheet_depth(wb, openpyxl, depth, det=None, source=None):
    """
    깊이 검증 — 검출 화소로 푼 3D 가 참값과 몇 mm 다른가.

    화소 오차(px)와 각도 오차(°)만으로는 정작 궁금한 것이 빠진다.
    평활도는 점별 깊이 오차가 그대로 결과이므로, 이 표가 곧 평활도를
    믿을 수 있는지의 근거다.
    """
    ws = wb.create_sheet("4.깊이검증(2단계)")
    ws.append(["항목", "값", "설명"])
    src_ko = {"detected": "검출 화소 (이미지에서 찾은 것 — 실장비와 같음)",
              "gt": "정답 화소 (raycast — 카메라가 완벽했다면의 상한)"}
    ws.append(["검측에 넣은 화소", src_ko.get(source, source or "-"),
               "이 조서의 모든 3D·각도·평활도 값이 여기서 나왔다"])
    if not depth:
        ws.append(["깊이 검증", "수행 못함",
                   "raycast 참값(xyz_world)이 있는 입력에서만 잴 수 있다"])
        _style(ws, openpyxl, widths=[24, 30, 78])
        return ws
    mm = (det or {}).get("mm_per_px_depth")
    rows = [
        ("비교 대상 점", f"{depth['n_points']:,}점",
         "검출점과 정답점은 같은 점이 아니다. 같은 선에서 스캔축 좌표가 "
         "같은 자리의 참값을 선형보간해 꺼내 비교했다"),
        ("", "", ""),
        ("깊이 치우침 (중앙값)", f"{depth['z_bias_mm']:+.2f} mm",
         "모든 점이 한쪽으로 밀린 몫. 좌표 규약·주점·기선이 어긋나면 "
         "생기며 소프트웨어로 없앨 수 있다"),
        ("깊이 산포 (σ)", f"{depth['z_noise_mm']:.2f} mm",
         "치우침을 뺀 나머지. 이것이 실제 점별 깊이 정밀도이고, "
         "평활도(±2mm 목표)가 잴 수 있는지를 여기서 가른다"
         + (f". 화소 1px = 깊이 {mm} mm" if mm else "")),
        ("깊이 RMS", f"{depth['z_rms_mm']:.2f} mm", "치우침 + 산포"),
        ("깊이 95 백분위", f"{depth['z_p95_mm']:.2f} mm",
         "스무 점 중 한 점은 이보다 더 틀린다"),
        ("", "", ""),
        ("3D 위치 오차 (중앙)", f"{depth['dist_med_mm']:.2f} mm",
         "깊이만이 아니라 X·Y 까지 포함한 점 사이 거리"),
        ("3D 위치 오차 (95%)", f"{depth['dist_p95_mm']:.2f} mm", ""),
    ]
    for a, b, c in rows:
        ws.append([a, b, c])
    _style(ws, openpyxl, widths=[24, 30, 78])

    fc = depth.get("개선예측_mm")
    if fc:
        ws.append([])
        h = ws.max_row + 1
        ws.append(["이 오차를 줄이는 방법", "예상 깊이 산포", "설명"])
        ws.append(["현재 입력", f"{fc['현재']:.2f} mm",
                   "지금 넣은 이미지에서 실제로 나온 값"])
        if "원해상도로 저장" in fc:
            ws.append([f"원해상도로 저장 (×{fc['해상도 배율']})",
                       f"{fc['원해상도로 저장']:.2f} mm",
                       "화면 캡처를 줄여서 저장하면 그 배율만큼 서브픽셀 "
                       "정밀도를 버린다. 센서 해상도 그대로 저장하기만 해도 "
                       "이만큼 좋아진다 — 촬영 설정이지 알고리즘이 아니다"])
        if "+ 안티에일리어싱" in fc:
            ws.append(["+ 안티에일리어싱", f"{fc['+ 안티에일리어싱']:.2f} mm",
                       "선 가장자리에 밝기 기울기가 생겨 중심을 화소 사이에서 "
                       "읽을 수 있다. 이진(0/255)으로 그려진 지금은 그 정보가 "
                       "아예 없다"])
        ws.append(["", "", ""])
        ws.append(["하드웨어 쪽 지렛대", "",
                   "σ_Z = σ_u·Z²/(f·b) 이므로 기선 b 를 늘리거나, 초점거리 "
                   "f 를 늘리거나, 측정거리 Z 를 줄이면 함께 좋아진다. "
                   "Z 는 제곱으로 들어가 가장 세다 — 2.7m 에서 1.5m 로 "
                   "다가서면 오차가 1/3 이 된다"])
        _style(ws, openpyxl, widths=[24, 30, 78], header_row=h)

    pl = depth.get("per_line_mm") or {}
    if pl:
        ws.append([])
        h = ws.max_row + 1
        ws.append(["선", "깊이 오차 중앙값(mm)", ""])
        for lid in sorted(pl, key=lambda k: (k[0], int(k[1:]))):
            ws.append([lid, pl[lid], ""])
        _style(ws, openpyxl, widths=[24, 30, 78], header_row=h)
    return ws


def _sheet_segmentation(wb, openpyxl, result, record, label_pixels=None,
                        seg_image_path=None):
    """색깔별로 무엇을 무엇으로 구분했는지."""
    ws = wb.create_sheet("5.세그멘테이션(3단계)")
    ws.append(["색", "클래스(코드)", "부재", "영역 수", "격자점 수",
               "화소 수", "검측 항목", "적용 기준(KCS)"])
    from openpyxl.styles import PatternFill

    agg = {}
    for r in result.get("regions", []):
        cls = r.get("class")
        a = agg.setdefault(cls, {"n": 0, "pts": 0, "kind": r.get("kind")})
        a["n"] += 1
        a["pts"] += int(r.get("n_points") or 0)

    seg = result.get("segmentation") or {}
    for cls in sorted(agg, key=lambda c: -agg[c]["pts"]):
        a = agg[cls]
        px = ""
        if label_pixels is not None:
            px = int(label_pixels.get(cls, 0))
        ws.append(["", cls, CLASS_KO.get(cls, cls), a["n"], a["pts"], px,
                   KIND_KO.get(a["kind"], a["kind"] or "—"),
                   KCS_NOTE.get(cls, "—")])
        c = ws.cell(row=ws.max_row, column=1)
        c.fill = PatternFill("solid",
                             fgColor=_rgb_hex(CLASS_COLOR.get(cls, (110,) * 3)))

    ws.append([])
    ws.append(["백엔드", seg.get("backend"), "", "", "", "", "", ""])
    cav = seg.get("caveat")
    if cav:
        ws.append(["한계", cav, "", "", "", "", "", ""])
    ws.append(["읽는 법",
               "색은 아래 세그멘테이션 이미지의 점 색과 같다. 같은 색이라도 "
               "적용 기준이 다를 수 있으므로 「적용 기준」 열을 함께 본다.",
               "", "", "", "", "", ""])
    _style(ws, openpyxl, widths=[6, 16, 12, 8, 11, 10, 12, 44])

    # 세그멘테이션 이미지를 시트에 붙인다. 색 범례와 그림이 떨어져 있으면
    # 어느 색이 무엇인지 대조하느라 시트를 오가야 한다.
    _embed_image(ws, seg_image_path, ws.max_row + 3, max_px=1500,
                 openpyxl=openpyxl,
                 caption="[그림] 세그멘테이션 결과 — 원본 사진 위에 부재별 "
                         "색으로 찍은 격자점, 자홍색 원이 요철 위치")
    return ws


def _sheet_results(wb, openpyxl, record):
    """구분별 품질검측 결과 — 이 조서의 본문."""
    ws = wb.create_sheet("6.검측결과")
    ws.append(["No", "부재", "검측 항목", "측정 각도(°)", "허용",
               "편차(mm)", "자세 판정",
               "직선자(m)", "처짐(mm)", "처짐 상한(mm)", "허용(mm)",
               "평활 판정", "요철 깊이(mm)", "요철 수",
               "측정거리(m)", "법선 σ(mm)", "격자점 수", "비고"])
    from openpyxl.styles import PatternFill, Font
    for m in record["members"]:
        if m["상태"] != "measured":
            ws.append([m["no"], m["부재"], m["검측항목"], "", "", "", "측정불가",
                       "", "", "", "", "해당없음", "", "", "", "",
                       m["점수"], m.get("사유") or ""])
            continue
        j = m.get("기준") or {}
        f = m.get("평활도") or {}
        u = m.get("불확실도") or {}
        allow = (f"±{j.get('허용_mm')}mm" if j.get("방식") == "mm"
                 else f"±{j.get('허용_deg')}°")
        ws.append([
            m["no"], m["부재"], m["검측항목"], m.get("측정각도_deg"), allow,
            j.get("편차_mm"), m.get("판정"),
            f.get("직선자_m"), f.get("처짐량_mm"), f.get("처짐량상한_mm"),
            f.get("허용_mm"), f.get("판정"),
            f.get("요철깊이_mm"), f.get("요철개수"),
            u.get("측정거리_m"), u.get("법선방향_sigma_mm"),
            m["점수"], f.get("비고") or j.get("비고") or ""])
    # 판정 열 색칠
    for col in (7, 12):
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            c = row[0]
            cf = _verdict_fill(openpyxl, c.value)
            if cf:
                c.fill = PatternFill("solid", fgColor=cf[0])
                c.font = Font(bold=True, color=cf[1], size=10)
    _style(ws, openpyxl,
           widths=[5, 11, 11, 12, 11, 10, 11, 10, 10, 12, 10, 14, 12, 8,
                   11, 11, 10, 34])
    return ws


def _sheet_flatness(wb, openpyxl, result):
    """KCS 직선자 길이별 상세. 3m 와 1m 는 허용치가 다르다."""
    ws = wb.create_sheet("7.평활도상세")
    ws.append(["No", "부재", "직선자 길이(m)", "실제 측정 구간(m)",
               "처짐(mm)", "처짐 상한(mm)", "허용(mm)", "허용 대비",
               "판정", "비고"])
    from openpyxl.styles import PatternFill, Font
    n = 0
    for i, r in enumerate(result.get("regions", [])):
        f = r.get("flatness") or {}
        kcs = f.get("kcs") or {}
        for c in kcs.get("checks", []):
            n += 1
            ratio = c.get("ratio")
            ws.append([i + 1, CLASS_KO.get(r.get("class"), r.get("class")),
                       c.get("length_m"), c.get("span_m"),
                       c.get("max_gap_mm"), c.get("upper_estimate_mm"),
                       c.get("tolerance_mm"),
                       (f"{ratio*100:.0f}%" if ratio is not None else ""),
                       kcs.get("judgement"), kcs.get("note") or ""])
            cc = ws.cell(row=ws.max_row, column=9)
            cf = _verdict_fill(openpyxl, cc.value)
            if cf:
                cc.fill = PatternFill("solid", fgColor=cf[0])
                cc.font = Font(bold=True, color=cf[1], size=10)
    if n == 0:
        ws.append(["", "평활도가 적용되는 면 영역이 없습니다.",
                   "", "", "", "", "", "", "", ""])
    _style(ws, openpyxl, widths=[5, 11, 14, 16, 11, 13, 11, 10, 16, 40])
    return ws


def _sheet_defects(wb, openpyxl, result, seg_image_path=None):
    """검출된 요철이 화면 어디에 있는가."""
    ws = wb.create_sheet("8.요철위치")
    ws.append(["No", "부재", "요철", "깊이(mm)", "크기(mm)",
               "화면 중심 u,v (px)", "화면 범위 u1,v1–u2,v2 (px)",
               "측정거리(m)", "구성 점수"])
    n = 0
    for i, r in enumerate(result.get("regions", [])):
        f = r.get("flatness") or {}
        for k, dd in enumerate(f.get("defects") or []):
            n += 1
            x0, y0, x1, y1 = dd["bbox_px"]
            ws.append([i + 1, CLASS_KO.get(r.get("class"), r.get("class")),
                       k + 1, dd["depth_mm"], dd["extent_mm"],
                       f"{dd['center_px'][0]:.0f}, {dd['center_px'][1]:.0f}",
                       f"{x0:.0f}, {y0:.0f} – {x1:.0f}, {y1:.0f}",
                       dd["z_m"], dd["n_points"]])
    if n == 0:
        ws.append(["", "검출된 요철 없음", "", "", "", "", "", "", ""])
    _style(ws, openpyxl, widths=[5, 11, 6, 11, 11, 20, 26, 11, 10])
    ws.append([])
    ws.append(["읽는 법", "화면 좌표는 센서 화소 기준이며, 세그멘테이션 "
               "이미지에 자홍색 원으로 같은 위치가 표시된다. 깊이는 평활값이 "
               "아니라 그 자리의 원시 잔차에서 잰 값이다 — 평활 창이 요철보다 "
               "넓으면 깊이를 깎기 때문이다."])
    _embed_image(ws, seg_image_path, ws.max_row + 3, max_px=1500,
                 openpyxl=openpyxl,
                 caption="[그림] 요철 위치 — 자홍색 원과 깊이 표기")
    return ws


def _sheet_caveats(wb, openpyxl, record, extra=None):
    ws = wb.create_sheet("11.유의사항")
    ws.append(["구분", "내용"])
    for c in record.get("caveats", []):
        ws.append(["측정", c])
    for c in (extra or []):
        ws.append(["설계값", c])
    prov = CALIB.provenance()
    assumed = [f"{k}: {v[1]}" for k, v in prov.items() if v[0] == "assumed"]
    for a in assumed:
        ws.append(["가정값", a])
    if ws.max_row == 1:
        ws.append(["", "특기사항 없음"])
    _style(ws, openpyxl, widths=[10, 100])
    return ws


# =====================================================================
# 진입점
# =====================================================================
def _sheet_pointcloud(wb, openpyxl, result, g_hat=None, image_path=None):
    """
    3D 좌표 — 부재별로 점이 어디에 찍혔는가 (요약).

    검측표는 "벽 수직도 0.03°" 같은 숫자만 준다. 그 숫자가 어느 점들에서
    나왔는지, 부재가 제대로 갈렸는지는 3D 배치를 봐야 안다. 여기에는
    부재별 점 수·중심·크기·거리 범위를 싣고, 좌표 원본은 다음 시트와
    CSV 로 낸다.
    """
    ws = wb.create_sheet("9.3D좌표(부재별)")
    rows = REPORT.region_xyz_summary(result, g_hat=g_hat)
    ws.append(["부재", "클래스", "검측", "상태", "격자선 수", "측정 점수",
               "가로선 점수", "선 구성", "깊이이득", "부재 반폭(mm)",
               "폭 근거", "중심 X(m)", "중심 Y(m)",
               "중심 Z(m)", "거리범위(m)", "가로폭(m)", "깊이폭(m)",
               "측정 구간(m)", "길이 확정?", "길이 근거"])
    for r in rows:
        # 크기를 어떻게 읽어야 하는지 한 칸에 적는다. 숫자만 두면
        # 거리 때문에 달라진 값을 부재 길이 차이로 오해한다.
        if r.get("선형부재"):
            how = ("격자가 닿은 구간만 잰 값 — 부재 길이가 **아니다**. "
                   "선형 부재는 끝이 화면·격자 밖에 있는 것이 보통이라, "
                   "거리가 다르면 같은 화소 구간이라도 미터값이 달라진다")
        elif r.get("구간제한"):
            how = ("격자가 닿은 끝까지 — 위 크기는 **하한**이다. 부재가 "
                   "거기서 끝난 게 아니라 측정 범위가 끝난 것이다")
        else:
            how = "부재 전체가 측정 범위 안"
        ws.append([
            r["부재번호"], r["클래스"], r["검측"], r["상태"],
            r.get("선수"), r["점수"], r.get("가로선점수", 0),
            r.get("선구성"), r.get("깊이이득_RMS"),
            r.get("부재반폭_mm"), r.get("폭_근거"),
            r["중심_X_m"], r["중심_Y_m"], r["중심_Z_m"],
            f"{r['거리범위_m'][0]} ~ {r['거리범위_m'][1]}",
            r.get("가로폭_m"), r.get("깊이폭_m"),
            # 하한이면 숫자 앞에 ≥ 를 붙인다. 표를 훑는 사람이 주석까지
            # 읽지 않아도 "이건 길이가 아니다" 가 눈에 들어와야 한다.
            ((f"≥ {r.get('측정구간_m')}" if r.get("길이확정") == "하한"
              else r.get("측정구간_m")) if r.get("측정구간_m") is not None
             else None),
            r.get("길이확정"), r.get("길이근거") or how])
    if not rows:
        ws.append(["3D 점이 없다 — 삼각측량 단계를 확인할 것"])
    _style(ws, openpyxl,
           widths=[7, 13, 14, 10, 10, 11, 11, 18, 9, 11, 11, 11, 17, 11, 11,
                   11, 46])

    n = ws.max_row + 2
    ws.cell(row=n, column=1, value="좌표계 설명").font = \
        openpyxl.styles.Font(bold=True)
    for i, t in enumerate([
        "X·Y·Z 는 조사기(레이저 원점) 좌표다. X=오른쪽, Y=아래, Z=정면 거리. "
        "eq1·eq7 삼각측량이 직접 낸 값이며 검측식이 쓴 것과 같은 배열이다.",
        "가로폭·깊이폭·높이폭은 중력 정렬 좌표에서 잰 외접 상자 크기다. "
        "장비를 숙이고 찍어도 벽은 수직, 바닥은 수평으로 서게 돌려 놓은 축이다.",
        "높이의 원점은 장비 광학중심이다. 절대 표고가 아니므로 표고로 바꾸려면 "
        "장비 설치 높이를 더해야 한다 — 그 값은 이 장비가 알 수 없다.",
        "선 구성은 그 부재의 점이 V선(세로)·H선(가로) 중 어디서 왔는지다. "
        "가로선이 깊이를 못 주는 사양에서는 V 만 나온다.",
        "격자선 수가 1 이면 그 부재는 레이저 평면 한 장에만 걸린 것이다. "
        "그때 나온 각도는 그 평면 안 성분일 뿐이라 참 기울기는 그보다 "
        "크거나 같고, 판정은 합격이 아니라 판정보류(단면 미분해)가 된다.",
        "「크기 해석」을 먼저 읽을 것. 동바리·철근 같은 선형 부재의 "
        "높이폭은 격자가 닿은 구간이지 부재 길이가 아니다. 실측에서 "
        "동바리 세 본이 1.99/1.93/1.87m 로 나왔는데 거리 1.705/1.652/1.595m "
        "에 정확히 비례한 값이었다 — 같은 화소 구간을 서로 다른 거리에서 "
        "본 것일 뿐 부재 길이가 다른 것이 아니다.",
        "가로선 점수는 깊이를 못 주는 선의 화소를 이미 맞춘 면 위에 올려 "
        "위치만 표시한 것이다. 화소 위치는 실제 검출값이지만 거리는 면에서 "
        "빌려 온 유도값이라, 그 면에 대한 잔차가 정의상 0 이다. 그래서 "
        "수직도·수평도·평활도 계산에는 넣지 않았다.",
    ]):
        ws.cell(row=n + 1 + i, column=1, value=t)
        ws.merge_cells(start_row=n + 1 + i, start_column=1,
                       end_row=n + 1 + i, end_column=17)
    _embed_image(ws, image_path, n + 6, max_px=1700, openpyxl=openpyxl,
                 caption="[그림] 부재별 3D 점군 — 등각·평면도·정면도·측면도. "
                         "중력 정렬 좌표라 벽은 서고 바닥은 눕는다")
    return ws


def _sheet_pointcloud_xyz(wb, openpyxl, result, g_hat=None, stride=20,
                          max_rows=50000):
    """
    3D 좌표 원본 — 부재 구분이 붙은 점 목록.

    stride 로 솎아 낸다. 3만 점을 그대로 넣으면 파일이 무거워지고 사람이
    읽지도 않는다. 전체 좌표는 같은 폴더의 CSV 에 있고, 앞 시트의 요약
    통계는 언제나 **전체 점** 으로 낸 값이다.
    """
    ws = wb.create_sheet("10.3D좌표(점목록)")
    rows = REPORT.region_xyz_rows(result, g_hat=g_hat, stride=stride)
    if len(rows) > max_rows:
        rows = rows[:max_rows]
        trimmed = True
    else:
        trimmed = False
    if not rows:
        ws.append(["3D 점이 없다"])
        _style(ws, openpyxl, widths=[40])
        return ws
    cols = list(rows[0].keys())
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    _style(ws, openpyxl, widths=[9, 13, 15, 8] + [11] * (len(cols) - 4))
    n = ws.max_row + 2
    note = (f"{stride}개마다 한 점씩 실었다. 전체 좌표는 같은 폴더의 "
            f"_3D좌표.csv 에 있다.")
    if trimmed:
        note += f" 시트 상한 {max_rows:,}행에서 잘랐다."
    ws.cell(row=n, column=1, value=note)
    ws.merge_cells(start_row=n, start_column=1, end_row=n,
                   end_column=len(cols))
    return ws


def save_excel(path, result, meta=None, label_pixels=None,
               extra_caveats=None, seg_image_path=None, detection=None,
               end_to_end=None, triangulation=None, g_hat=None,
               pointcloud_image=None, pc_stride=20, detection_image=None,
               depth_check=None, pixel_source=None):
    """
    검측 결과를 엑셀 조서로 저장한다.

    Parameters
    ----------
    result       : pipeline_region.inspect_image / inspect_capture 결과
    meta         : dict — 입력 이미지, 촬영 조건 등 (요약 시트 상단)
    label_pixels : {class: 화소수} — 세그멘테이션 마스크 면적 (선택)
    seg_image_path : 세그멘테이션 결과 이미지 경로 — 4번 시트에 삽입
    detection    : load_capture.evaluate_line_detection 결과 (선택)
    g_hat        : (3,) 조사기 좌표계 중력. 주면 3D 좌표 시트가 중력 정렬
                   좌표(가로·깊이·높이)를 함께 낸다
    pointcloud_image : 3D 점군 이미지 경로 — 3D 좌표 시트에 삽입
    detection_image : 선검출 대조 이미지 경로 — 선검출정확도 시트에 삽입
    depth_check  : load_capture.verify_depth 결과 — 깊이검증 시트
    pixel_source : "detected" | "gt" — 검측에 넣은 화소가 무엇이었는지
    pc_stride    : 3D 점목록 시트에 N개마다 한 점 (기본 20)
    """
    import openpyxl
    record = REPORT.build_record(result, meta)
    wb = openpyxl.Workbook()
    _sheet_summary(wb, openpyxl, record, dict(meta or {}), result)
    _sheet_design(wb, openpyxl)
    _sheet_detection(wb, openpyxl, detection, end_to_end, triangulation,
                     overlay_image=detection_image)
    _sheet_depth(wb, openpyxl, depth_check, detection, pixel_source)
    _sheet_segmentation(wb, openpyxl, result, record, label_pixels,
                        seg_image_path)
    _sheet_pointcloud(wb, openpyxl, result, g_hat=g_hat,
                      image_path=pointcloud_image)
    _sheet_results(wb, openpyxl, record)
    _sheet_flatness(wb, openpyxl, result)
    _sheet_defects(wb, openpyxl, result, seg_image_path)
    _sheet_pointcloud_xyz(wb, openpyxl, result, g_hat=g_hat, stride=pc_stride)
    # '동바리 높이가 제각각' 으로 읽히는 것을 막는 설명은 조서에도 남긴다.
    _sn = (result.get("summary") or {}).get("span_note")
    _cav = list(extra_caveats or [])
    if _sn:
        _cav.insert(0, _sn)
    _sheet_caveats(wb, openpyxl, record, _cav)
    # 탭 순서를 파이프라인 순서로 맞춘다 — 선검출 → 깊이 → 분할/3D → 검측
    order = ["1.요약", "2.설계값", "3.선검출(1단계)", "4.깊이검증(2단계)",
             "5.세그멘테이션(3단계)", "6.검측결과", "7.평활도상세",
             "8.요철위치", "9.3D좌표(부재별)", "10.3D좌표(점목록)",
             "11.유의사항"]
    have = [t for t in order if t in wb.sheetnames]
    wb._sheets = ([wb[t] for t in have]
                  + [w for w in wb.worksheets if w.title not in have])
    d = _os.path.dirname(_os.path.abspath(path))
    if d:
        _os.makedirs(d, exist_ok=True)
    wb.save(path)
    return path
