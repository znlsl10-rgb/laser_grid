#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
experiment.py — [실험용] 기선거리 × 측정거리 sweep → sweep_results.json
========================================================================
inspection.py의 A,B 알고리즘을 그대로 사용하여
기선(baseline) × 측정거리(standoff) 30조합에서
픽셀 검출 정확도(pixel_rmse_px)와 깊이 정확도(sigma_z_mm)를 측정한다.

A,B 알고리즘 교체:
  python3 experiment.py \
    --algo-detect     새_A알고리즘.py \
    --algo-triangulate 새_B알고리즘.py

출력 (→ experiment_compare.py 입력):
  experiment_output/sweep_results.json

실행:
  python3 experiment.py
========================================================================
"""
import os, sys, json, argparse, importlib.util
import numpy as np
from PIL import Image, ImageDraw

# =====================================================================
# inspection.py 공용 설정 import
# =====================================================================
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from inspection import (
    CAMERA_PARAMS, GRID_PARAMS, STATIONS,
    capture_station,
    SCENE_USD, GT_JSON,
    _norm, _wait,
    _setup_scene, _setup_camera, _make_line_angles,
    fn_detect as _default_detect,
    LOG, _RUNNING_IN_GUI,
)

# B 삼각측량 직접 로드
import importlib.util as _ilu, os as _os
def _load_tri(path):
    spec = _ilu.spec_from_file_location("_B", path)
    m = _ilu.module_from_spec(spec); spec.loader.exec_module(m)
    return m.triangulate
_B_PATH = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                         "B_삼각측량.py")
try:
    _default_triangulate = _load_tri(_B_PATH)
except Exception as _e:
    print(f"[경고] B_삼각측량 로드 실패: {_e}")
    _default_triangulate = None

OUTPUT_DIR = "/home/develop/Desktop/laser_grid_test_4/result/experiment_output"

# =====================================================================
# 실험 파라미터
# =====================================================================
# =====================================================================
# 실험 파라미터 — 검측 항목 × 조작 변수 전체
# =====================================================================
# 수직도/수평도  : 1-1_build에서 HORIZ=0.3°, VBACK=0.5° 고정 (한 세트)
# 거리 측정(Z)   : standoffs_m    [0.5, 1.0, 1.5, 2.0, 3.0, 4.0] m
# 기선 (설계검증): baselines_mm   [80, 100, 120, 150, 180, 200] mm
SWEEP_PARAMS = {
    # ── 기선 변화 (설계값 150mm 검증) ──
    "baselines_mm":       [80, 100, 120, 150, 180, 200],

    # ── 거리 측정 (Z) 6단계 ──
    "standoffs_m":        [0.5, 1.0, 1.5, 2.0, 3.0, 4.0],

    # ── 수직도 벽 기울기 6단계 (°) ──
    "wall_tilts_deg":     [0, 1, 2, 3, 5, 10],

    # ── 수평도 바닥 기울기 5단계 (°) ──
    "floor_tilts_deg":    [0, 1, 2, 3, 5],

    # ── 평활도 요철 진폭 4단계 (mm) ──
    "bump_amps_mm":       [0, 2, 5, 10],

    # ── IMU 기울기 (고정, 핸드헬드 모사) ──
    "imu_pitch_deg":      1.0,
    "imu_roll_deg":       0.5,

    # ── 목표 정확도 ──
    "target_sigma_z_mm":  2.0,

    # ── 기본 실험 스테이션 ──
    "station":            "StationA_Wall",
}

# =====================================================================
# 검증1: 픽셀 검출 정확도
# =====================================================================
def compute_pixel_accuracy(lines_detected, lines_raycast, line_angles):
    """
    [검증1] V/H 선 전체에 대해 point-wise 픽셀 오차 RMSE.

    v5 수정: 기존 방식(선 1개당 median 오차값 1개)은 선 내부의 국소
    오차(일부 구간 점프, 결함부 둔감화 등)를 전혀 반영하지 못하고,
    V선만 계산해 H선 오차가 지표에서 누락되어 있었다(line_angles의
    fixed=="alpha" 조건으로 H선을 건너뜀).

    compute_depth_accuracy와 동일한 방식으로, GT(raycast, 보통 250샘플)와
    검출 결과(가변 개수)의 샘플링 밀도가 다르므로 인덱스 매칭이 아니라
    독립변수(축) 기준 선형보간으로 GT를 검출점 위치에 맞춰 재추출한 뒤
    그 차이를 point-wise로 비교한다.
    V선은 행(v)을 따라가며 u오차를, H선은 열(u)을 따라가며 v오차를 본다.
    """
    errs=[]; v_errs=[]; h_errs=[]; per_line={}; n_total=0; n_hit=0
    for lid,det in lines_detected.items():
        ray=lines_raycast.get(lid,[])
        n_total+=1
        if not det or not ray or len(ray)<2:
            per_line[lid]={"rmse_px":None,"n":0}; continue

        is_v = lid.startswith("V")
        coord_idx = 0 if is_v else 1   # 오차를 잴 좌표(V선=u, H선=v)
        axis_idx  = 1 if is_v else 0   # 보간 기준 독립변수(V선=v행, H선=u열)

        det_arr = np.array(det, dtype=float)
        ray_arr = np.array(ray, dtype=float)
        order = np.argsort(ray_arr[:, axis_idx])
        ray_arr = ray_arr[order]
        ray_axis  = ray_arr[:, axis_idx]
        ray_coord = ray_arr[:, coord_idx]
        if len(np.unique(ray_axis)) < 2:
            per_line[lid]={"rmse_px":None,"n":0}; continue

        det_axis  = det_arr[:, axis_idx]
        det_coord = det_arr[:, coord_idx]
        gt_interp = np.interp(det_axis, ray_axis, ray_coord)

        line_errs = np.abs(det_coord - gt_interp)
        errs.extend(line_errs.tolist())
        (v_errs if is_v else h_errs).extend(line_errs.tolist())
        per_line[lid] = {"rmse_px": round(float(np.sqrt(np.mean(line_errs**2))), 3),
                          "n": int(len(line_errs))}
        n_hit += 1

    def _rmse_of(a):
        return round(float(np.sqrt(np.mean(np.array(a)**2))), 4) if a else None

    errs_arr = np.array(errs) if errs else None
    rmse = float(np.sqrt(np.mean(errs_arr**2))) if errs else None
    return {"pixel_rmse_px": round(rmse,4) if rmse else None,
            "pixel_mean_px": round(float(np.mean(errs_arr)),4) if errs else None,
            "pixel_max_px":  round(float(np.max(errs_arr)),4)  if errs else None,
            "v_rmse_px":     _rmse_of(v_errs),
            "h_rmse_px":     _rmse_of(h_errs),
            "hit_ratio":     round(n_hit/max(n_total,1),3),
            "n_compared":    int(len(errs)),
            "per_line":      per_line}

# =====================================================================
# 검증2: 깊이 정확도
# =====================================================================
def compute_depth_accuracy(depth_points, ground_truth):
    """
    [검증2] Z_rec vs z_cam_m → σZ.

    depth_points와 ground_truth는 서로 다른 그리드(raycast 250샘플 vs
    A검출 가변 개수)이므로, 단순 인덱스 매칭이 아니라
    v좌표(행) 기준으로 가장 가까운 GT를 찾아 비교한다.
    """
    from collections import defaultdict
    by_lid=defaultdict(list)
    for dp in depth_points: by_lid[dp["lid"]].append(dp)
    z_errs=[]; per_line={}
    for lid,dp_list in by_lid.items():
        gt_pts=ground_truth.get(lid,[])
        if not gt_pts: per_line[lid]={"z_rmse_mm":None,"n":0}; continue

        # GT를 v(또는 u, H선의 경우)좌표 기준 정렬된 배열로 준비
        gt_arr = [(g.get("v_px", g.get("u_px", i)), g.get("z_cam_m"))
                  for i, g in enumerate(gt_pts) if "z_cam_m" in g]
        if not gt_arr: per_line[lid]={"z_rmse_mm":None,"n":0}; continue
        gt_arr.sort(key=lambda x: x[0])
        gt_coords = np.array([g[0] for g in gt_arr])
        gt_zs     = np.array([g[1] for g in gt_arr])

        le=[]
        for dp in dp_list:
            v_dp = dp.get("v_px")
            if v_dp is None: continue
            # 가장 가까운 GT 좌표 탐색 (선형 보간)
            z_gt = float(np.interp(v_dp, gt_coords, gt_zs))
            err = abs(float(dp["Z_rec"]) - z_gt) * 1000
            le.append(err); z_errs.append(err)
        per_line[lid]={"z_rmse_mm":_rmse(le),"n":len(le)}
    return {"sigma_z_mm":   _rmse(z_errs),
            "mean_z_err_mm":round(float(np.mean(z_errs)),4) if z_errs else None,
            "max_z_err_mm": round(float(np.max(z_errs)),4)  if z_errs else None,
            "n_compared":   len(z_errs), "per_line": per_line}

def _rmse(arr):
    if not arr: return None
    return round(float(np.sqrt(np.mean(np.array(arr)**2))),4)

# =====================================================================
# 단일 조건 실험
# =====================================================================
def _apply_scene_tilt(stage, target, deg):
    """
    Wall 또는 Floor prim의 x축 회전을 변경한다.

    1-1_build_inspection_lab_realistic.py 원본 회전 방식 그대로 사용
    (메시 로컬중심(2.5,1.5,0)이 rotate_xyz(90+deg,0,0)에서
     약 (2.5, 0, 1.5) 근처로 매핑되도록 설계됨 — 피벗보정 불필요):
      WallBackFace: rotate_xyz = (90 + VBACK + deg, 0, 0)
      FloorTop:      rotate_xyz = (HORIZ + deg, 0, 0)
    """
    try:
        from pxr import UsdGeom, Gf
        BASE_DEG = {"wall": 90.5, "floor": 0.3}   # 90+VBACK, HORIZ
        paths = {"wall": "/World/StationA/WallBackFace",
                 "floor": "/World/StationA/FloorTop"}
        p = stage.GetPrimAtPath(paths.get(target,""))
        if not p or not p.IsValid(): return
        total_deg = BASE_DEG.get(target, 0.0) + float(deg)
        xf = UsdGeom.Xformable(p)
        xf.ClearXformOpOrder()
        xf.AddRotateXYZOp().Set(Gf.Vec3f(total_deg, 0., 0.))
    except Exception as e:
        LOG(f"  [경고] 기울기 적용 실패: {e}")


def _apply_bump_amplitude(stage, amp_mm):
    """StationB 패널의 요철 진폭을 변경한다."""
    try:
        from pxr import Sdf
        panel = stage.GetPrimAtPath("/World/StationB/Panel")
        if not panel or not panel.IsValid(): return
        attr = panel.GetAttribute("displacement:scale")
        if not attr:
            attr = panel.CreateAttribute("displacement:scale",
                                         Sdf.ValueTypeNames.Float)
        attr.Set(float(amp_mm) / 1000.)
    except Exception as e:
        LOG(f"  [경고] 요철 진폭 적용 실패: {e}")


# ── eq2/eq3/eq4 동적 로드 (없으면 None) ─────────────────────────────
def _try_import(name):
    try:
        spec = importlib.util.spec_from_file_location(
            name, os.path.join(os.path.dirname(os.path.abspath(__file__)), f"{name}.py"))
        m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m)
        return m
    except Exception as _e:
        LOG(f"  [품질검측] {name}.py 로드 실패: {_e}")
        return None

_EQ2 = _try_import("eq2_plane_fit")
_EQ3 = _try_import("eq3_orientation")
_EQ4 = _try_import("eq4_flatness_line")


def _run_quality_metrics(depth_points, extra):
    """
    삼각측량 결과(depth_points)로 eq2/eq3/eq4 품질검측 수행.

    Returns dict:
      verticality_deg   - 수직도 각도 [°]
      horizontality_deg - 수평도 각도 [°]
      flatness_max_mm   - 평활도 최대 잔차 [mm]
      flatness_rms_mm   - 평활도 RMS 잔차 [mm]
      flatness_pass     - 평활도 합격 여부 (bool)
      defect_clusters   - 요철 클러스터 수
      plane_normal      - 법선 벡터 (a,b,c)
      orientation_pass  - 수직도/수평도 허용 오차 이내 (bool)
    """
    if not depth_points or _EQ2 is None or _EQ3 is None:
        return {}

    exp_type = (extra or {}).get("exp_type", "")

    # 3D 점군 구성
    pts = np.array([[d["X_rec"] if "X_rec" in d else
                     (d.get("u_px",0)-CAMERA_PARAMS["cx_px"])*d["Z_rec"]
                        /CAMERA_PARAMS["f_px"]+CAMERA_PARAMS["b_m"],
                     (d.get("v_px",0)-CAMERA_PARAMS["cy_px"])*d["Z_rec"]
                        /CAMERA_PARAMS["f_px"],
                     d["Z_rec"]]
                    for d in depth_points if d.get("Z_rec") and d["Z_rec"]>0],
                   dtype=float)

    if len(pts) < 10:
        return {}

    result = {}
    try:
        plane, _ = _EQ2.fit_plane_ransac(pts, threshold=0.01)
        result["plane_normal"] = [round(float(x),6) for x in plane[:3]]

        # 수직도·수평도
        if exp_type == "verticality":
            theta = _EQ3.measure_verticality(plane[:3])
            result["verticality_deg"]  = round(theta, 4)
            result["orientation_pass"] = bool(abs(theta) <= 0.5)
        elif exp_type == "horizontality":
            theta = _EQ3.measure_horizontality(plane[:3])
            result["horizontality_deg"] = round(theta, 4)
            result["orientation_pass"]  = bool(abs(theta) <= 0.5)
        else:
            result["verticality_deg"]   = round(_EQ3.measure_verticality(plane[:3]), 4)
            result["horizontality_deg"] = round(_EQ3.measure_horizontality(plane[:3]), 4)

        # 평활도
        if _EQ4 is not None:
            fd = _EQ4.detect_defects_grid(pts, threshold_mm=2.0, grid_n=21)
            con = _EQ4.reconstruct_defect_contour(fd)
            result["flatness_max_mm"]  = round(fd["overall_max_dev_mm"], 3)
            result["flatness_rms_mm"]  = round(fd["rms_dev_mm"], 3)
            result["flatness_pass"]    = bool(fd["is_pass"])
            result["defect_clusters"]  = int(con["n_clusters"])
            result["defect_depth_mm"]  = round(con["depth_mm"], 3)

    except Exception as _e:
        LOG(f"  [품질검측 오류] {_e}")

    return result


def run_sweep(stage, world, camera, line_angles, fn_detect, fn_triangulate, gt_full,
             full_grid=False, use_diff_image=False):
    """
    검측 항목별 독립 실험 수행 (독립 변수별 단일 변경).

      [실험1] 거리 변화      : 0.5~4.0m  6단계
      [실험2] 수직도 기울기  : 0~10°     6단계   Wall x축 회전
      [실험3] 수평도 기울기  : 0~5°      5단계   Floor x축 회전
      [실험4] 평활도 요철    : 0~10mm    4단계   Displacement map
      [실험5] 기선 변화      : 80~200mm  6단계
      [실험6] 기선×거리 전체격자(옵션) : 6×6=36조합, full_grid=True일 때만 실행
              (실험1/5는 십자형 단면만 보므로, 진짜 최적 (기선,거리)
               조합을 찾으려면 전체 격자가 필요함)
    """
    sp = SWEEP_PARAMS
    all_results = []
    _gt_full = gt_full
    B_DEFAULT = 150; D_DEFAULT = 1.0; N_DEFAULT = 0.0
    _orig_station = sp["station"]

    def _run(label, b_mm, d_m, noise=0.0, extra=None):
        LOG(f"  ▶ {label}")
        # inspection.py의 capture_station() 직접 호출
        # baseline_m, standoff_m, noise_sigma_px 오버라이드
        cfg_now = dict(STATIONS[sp["station"]])
        res = capture_station(
            stage, world, camera, line_angles,
            station_name   = label,
            cfg            = cfg_now,
            gt_full        = _gt_full,
            baseline_m     = b_mm / 1000.,
            standoff_m     = d_m,
            noise_sigma_px = noise,
            output_dir     = OUTPUT_DIR,
            use_diff_image = use_diff_image,
        )
        if res:
            # 검증1: 픽셀 정확도, 검증2: 깊이 정확도 계산
            pa = compute_pixel_accuracy(
                res["lines_pixels"], res["lines_pixels_raw"], line_angles)
            # B 삼각측량으로 depth_points 생성
            cp_now = res.get("camera_params", {})
            b_now  = b_mm / 1000.
            la     = res.get("line_angles", line_angles)
            depth_points = []
            for lid, pts in res.get("lines_pixels", {}).items():
                if not la.get(lid, {}).get("fixed") == "alpha": continue
                alpha = la[lid]["angle_rad"]
                for u_v in pts:
                    u_px, v_px = float(u_v[0]), float(u_v[1])
                    Z = fn_triangulate(u_px, v_px, alpha, cp_now, b_now)
                    if Z:
                        depth_points.append({"lid":lid,"u_px":u_px,
                                             "v_px":v_px,"Z_rec":float(Z)})
            da = compute_depth_accuracy(depth_points, res["ground_truth"])

            # ── eq2/eq3/eq4 품질검측 연동 ──────────────────────────────
            qm = _run_quality_metrics(depth_points, extra)

            res["pixel_accuracy"]   = pa
            res["depth_accuracy"]   = da
            res["quality_metrics"]  = qm
            res["baseline_mm"]      = round(b_mm, 1)
            res["standoff_m"]       = d_m
            if extra: res.update(extra)
            res["experiment"] = label
            all_results.append(res)
            LOG(f"    픽셀RMSE={pa['pixel_rmse_px']}px  σZ={da['sigma_z_mm']}mm")
        else:
            LOG(f"    → 실패")

    total = (len(sp["standoffs_m"]) + len(sp["wall_tilts_deg"]) +
             len(sp["floor_tilts_deg"]) + len(sp["bump_amps_mm"]) +
             len(sp["baselines_mm"]))
    if full_grid:
        total += (len(sp["baselines_mm"]) * len(sp["standoffs_m"])
                  - len(sp["baselines_mm"]) - len(sp["standoffs_m"]) + 1)
    LOG(f"\n{'='*65}")
    LOG(f"[실험] 검측항목 전체  총 {total}조합")
    LOG(f"  [A] {getattr(fn_detect,'__module__','A_선검출')}")
    LOG(f"  [B] {getattr(fn_triangulate,'__module__','B_삼각측량')}")
    LOG(f"{'='*65}")

    # [실험1] 거리 변화
    LOG(f"\n[실험1] 거리 변화  b={B_DEFAULT}mm")
    sp["station"] = "StationA_Wall"
    for d in sp["standoffs_m"]:
        _run(f"dist_{d}m", B_DEFAULT, d, N_DEFAULT,
             extra={"exp_type":"distance"})

    # [실험2] 수직도 (Wall x축 회전)
    LOG(f"\n[실험2] 수직도 기울기  d={D_DEFAULT}m")
    sp["station"] = "StationA_Wall"
    for tilt in sp["wall_tilts_deg"]:
        _apply_scene_tilt(stage, "wall", tilt)
        _wait(world, 10)
        _run(f"wall_{tilt}deg", B_DEFAULT, D_DEFAULT, N_DEFAULT,
             extra={"exp_type":"verticality","wall_tilt_deg":tilt})
    _apply_scene_tilt(stage, "wall", 0.5)   # 원본 VBACK=0.5° 복원

    # [실험3] 수평도 (Floor x축 회전)
    LOG(f"\n[실험3] 수평도 기울기  d={D_DEFAULT}m")
    sp["station"] = "StationA_Floor"
    for tilt in sp["floor_tilts_deg"]:
        _apply_scene_tilt(stage, "floor", tilt)
        _wait(world, 10)
        _run(f"floor_{tilt}deg", B_DEFAULT, D_DEFAULT, N_DEFAULT,
             extra={"exp_type":"horizontality","floor_tilt_deg":tilt})
    _apply_scene_tilt(stage, "floor", 0.3)  # 원본 HORIZ=0.3° 복원

    # [실험4] 평활도 (Displacement map)
    LOG(f"\n[실험4] 평활도 요철  d={D_DEFAULT}m")
    sp["station"] = "StationB"
    for amp in sp["bump_amps_mm"]:
        _apply_bump_amplitude(stage, amp)
        _wait(world, 10)
        _run(f"bump_{amp}mm", B_DEFAULT, D_DEFAULT, N_DEFAULT,
             extra={"exp_type":"flatness","bump_amp_mm":amp})
    _apply_bump_amplitude(stage, 0)

    # [실험5] 기선 변화
    LOG(f"\n[실험5] 기선 변화  d={D_DEFAULT}m")
    sp["station"] = "StationA_Wall"
    for b_mm in sp["baselines_mm"]:
        _run(f"baseline_{b_mm}mm", b_mm, D_DEFAULT, N_DEFAULT,
             extra={"exp_type":"baseline"})

    # [실험6] 기선×거리 전체 격자 (옵션, --full-grid 지정 시에만)
    # 실험1/5는 십자형 단면(서로 다른 1개 축만 변경)이라 baseline×거리의
    # 진짜 조합 효과(예: 가까운 거리+큰 baseline)를 알 수 없다.
    # 이미 측정된 교차점(d=1.0×모든b, b=150×모든d)은 다시 돌리지 않고
    # 건너뛰어 중복 실행을 줄인다.
    if full_grid:
        n_grid = len(sp["baselines_mm"]) * len(sp["standoffs_m"])
        LOG(f"\n[실험6] 기선×거리 전체격자  {n_grid}조합(교차점 제외)")
        sp["station"] = "StationA_Wall"
        for b_mm in sp["baselines_mm"]:
            for d in sp["standoffs_m"]:
                if b_mm == B_DEFAULT and d == D_DEFAULT:
                    continue   # 실험1/5에서 이미 측정됨
                if b_mm == B_DEFAULT:
                    continue   # 실험1에서 이미 측정됨 (모든 d, b=150)
                if d == D_DEFAULT:
                    continue   # 실험5에서 이미 측정됨 (모든 b, d=1.0)
                _run(f"grid_b{b_mm}mm_d{d}m", b_mm, d, N_DEFAULT,
                     extra={"exp_type":"grid"})

    sp["station"] = _orig_station

    LOG(f"\n{'='*65}")
    LOG(f"총 {len(all_results)}/{total}조합 완료")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    def _s(o):
        if isinstance(o, float) and o!=o: return None
        if isinstance(o, np.floating): return float(o)
        if isinstance(o, np.integer):  return int(o)
        return o

    # ── JSON 저장 ──
    out_path = os.path.join(OUTPUT_DIR, "sweep_results.json")
    with open(out_path, "w", encoding="utf-8") as fp:
        json.dump(all_results, fp, ensure_ascii=False, indent=2, default=_s)
    LOG(f"JSON 저장 → {out_path}")

    # ── 엑셀 저장 ──
    _save_excel(all_results, OUTPUT_DIR)

    LOG("다음: python3 experiment_compare.py")
    return all_results


def _save_excel(all_results, out_dir):
    """실험 결과를 엑셀로 저장. 시트: 요약 / 실험별 상세 / 이미지."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        LOG("  [경고] openpyxl 없음 → 엑셀 저장 생략 (pip install openpyxl)")
        return

    import copy as _copy
    wb = openpyxl.Workbook()

    # ── 색상 정의 ──
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(color="FFFFFF", bold=True)
    PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
    FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
    ALT_FILL  = PatternFill("solid", fgColor="EEF3F8")
    CENTER    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _copy.copy(HDR_FILL); c.font = _copy.copy(HDR_FONT)
        c.alignment = CENTER; c.border = BORDER
        if width: ws.column_dimensions[get_column_letter(col)].width = width

    def _cell(ws, row, col, val, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = CENTER; c.border = BORDER
        if fill: c.fill = _copy.copy(fill)

    # ══════════════════════════════════════════
    # 시트1: 요약 (실험별 σZ, pixel_rmse)
    # ══════════════════════════════════════════
    ws1 = wb.active; ws1.title = "요약"
    ws1.row_dimensions[1].height = 22
    ws1.freeze_panes = "A2"

    headers = ["실험", "조건", "기선(mm)", "거리(m)",
               "pixel_rmse(px)", "σZ(mm)", "평균오차(mm)", "최대오차(mm)",
               "n_hit", "판정"]
    widths  = [18, 20, 10, 8, 16, 10, 12, 12, 8, 8]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws1, 1, ci, h, w)

    for ri, res in enumerate(all_results, 2):
        exp   = res.get("experiment", "")
        pa    = res.get("pixel_accuracy", {})
        da    = res.get("depth_accuracy", {})
        px    = pa.get("pixel_rmse_px")
        sz    = da.get("sigma_z_mm")
        mean_err = da.get("mean_z_err_mm")
        max_err  = da.get("max_z_err_mm")
        b_mm  = res.get("baseline_mm", "")
        d_m   = res.get("standoff_m", "")
        n_hit = res.get("n_hit", "")
        exp_type = res.get("exp_type", "")
        ok    = (sz is not None and sz <= 2.0)
        fill  = PASS_FILL if ok else FAIL_FILL if sz is not None else None
        row_fill = ALT_FILL if ri % 2 == 0 else None

        for ci, val in enumerate([exp, exp_type, b_mm, d_m,
                                   round(px,4) if px else None,
                                   round(sz,4) if sz else None,
                                   round(mean_err,4) if mean_err else None,
                                   round(max_err,4) if max_err else None,
                                   n_hit,
                                   "✅PASS" if ok else ("❌FAIL" if sz else "-")], 1):
            _cell(ws1, ri, ci, val, fill if ci in [6,10] else row_fill)

    # ══════════════════════════════════════════
    # 시트2~6: 실험별 상세 (exp_type별)
    # ══════════════════════════════════════════
    exp_types = {}
    for res in all_results:
        et = res.get("exp_type", "기타")
        exp_types.setdefault(et, []).append(res)

    type_names = {
        "distance":     "실험1_거리변화",
        "verticality":  "실험2_수직도",
        "horizontality":"실험3_수평도",
        "flatness":     "실험4_평활도",
        "baseline":     "실험5_기선변화",
    }

    for et, results in exp_types.items():
        sname = type_names.get(et, et)[:31]
        ws = wb.create_sheet(sname)
        ws.freeze_panes = "A2"

        hdrs = ["실험라벨", "기선(mm)", "거리(m)",
                "pixel_rmse(px)", "pixel_mean(px)", "hit_ratio",
                "σZ(mm)", "mean_z_err(mm)", "max_z_err(mm)",
                "n_compared", "n_hit", "n_valid_V", "판정"]
        ws_w = [22,10,8,16,16,10,10,14,14,12,8,10,8]
        for ci,(h,w) in enumerate(zip(hdrs,ws_w),1):
            _hdr(ws,1,ci,h,w)

        for ri, res in enumerate(results, 2):
            pa = res.get("pixel_accuracy",{})
            da = res.get("depth_accuracy",{})
            sz = da.get("sigma_z_mm")
            ok = sz is not None and sz <= 2.0
            vals = [
                res.get("experiment",""),
                res.get("baseline_mm",""),
                res.get("standoff_m",""),
                pa.get("pixel_rmse_px"),
                pa.get("pixel_mean_px"),
                pa.get("hit_ratio"),
                sz,
                da.get("mean_z_err_mm"),
                da.get("max_z_err_mm"),
                da.get("n_compared"),
                res.get("n_hit",""),
                res.get("n_valid_V",""),
                "✅PASS" if ok else ("❌FAIL" if sz else "-"),
            ]
            for ci,val in enumerate(vals,1):
                fill = PASS_FILL if (ci==7 and ok) else                        FAIL_FILL if (ci==7 and sz) else                        ALT_FILL  if ri%2==0 else None
                c = ws.cell(row=ri,column=ci,
                            value=round(val,4) if isinstance(val,float) else val)
                c.alignment=CENTER; c.border=BORDER
                if fill: c.fill=_copy.copy(fill)

    # ══════════════════════════════════════════
    # 시트6: 품질검측 결과 (eq2+eq3+eq4 통합)
    # ══════════════════════════════════════════
    ws_qm = wb.create_sheet("품질검측결과")
    ws_qm.freeze_panes = "A2"
    qm_hdrs = [
        "실험라벨", "exp_type", "기선(mm)", "거리(m)",
        "수직도(°)", "수평도(°)", "방향판정",
        "평활도최대(mm)", "평활도RMS(mm)", "평활도판정", "요철클러스터",
        "법선_a", "법선_b", "법선_c",
    ]
    qm_widths = [22,16,10,8,10,10,10,14,14,10,12,10,10,10]
    for ci,(h,w) in enumerate(zip(qm_hdrs,qm_widths),1):
        _hdr(ws_qm,1,ci,h,w)

    for ri, res in enumerate(all_results, 2):
        qm  = res.get("quality_metrics", {})
        exp = res.get("experiment","")
        et  = res.get("exp_type","")
        b   = res.get("baseline_mm","")
        d   = res.get("standoff_m","")
        vt  = qm.get("verticality_deg")
        ht  = qm.get("horizontality_deg")
        op  = qm.get("orientation_pass")
        fm  = qm.get("flatness_max_mm")
        fr  = qm.get("flatness_rms_mm")
        fp  = qm.get("flatness_pass")
        dc  = qm.get("defect_clusters")
        nn  = qm.get("plane_normal", [None,None,None])

        op_str = ("✅" if op else "❌") if op is not None else "-"
        fp_str = ("✅평탄" if fp else "❌요철") if fp is not None else "-"
        row = [exp, et, b, d,
               round(vt,4) if vt is not None else "-",
               round(ht,4) if ht is not None else "-",
               op_str,
               round(fm,3) if fm is not None else "-",
               round(fr,3) if fr is not None else "-",
               fp_str, dc if dc is not None else "-",
               round(nn[0],4) if nn[0] is not None else "-",
               round(nn[1],4) if nn[1] is not None else "-",
               round(nn[2],4) if nn[2] is not None else "-",]
        alt = ALT_FILL if ri%2==0 else None
        for ci,val in enumerate(row,1):
            fill = alt
            if ci==7:  fill = PASS_FILL if op else (FAIL_FILL if op is not None else alt)
            if ci==10: fill = PASS_FILL if fp else (FAIL_FILL if fp is not None else alt)
            _cell(ws_qm,ri,ci,val,fill)

    # ── 수직도 요약 소시트 ──
    vert_res = [r for r in all_results if r.get("exp_type")=="verticality"]
    if vert_res:
        ws_qm.cell(len(all_results)+4,1,"[수직도 요약]").font=_copy.copy(HDR_FONT)
        ws_qm.cell(len(all_results)+5,1,"GT 기울기(°)")
        ws_qm.cell(len(all_results)+5,2,"eq3 측정(°)")
        ws_qm.cell(len(all_results)+5,3,"오차(°)")
        ws_qm.cell(len(all_results)+5,4,"판정")
        for k,r in enumerate(sorted(vert_res,key=lambda x:x.get("wall_tilt_deg",0))):
            gt  = r.get("wall_tilt_deg",0)
            meas= r.get("quality_metrics",{}).get("verticality_deg")
            row = len(all_results)+6+k
            ws_qm.cell(row,1,gt)
            ws_qm.cell(row,2,round(meas,4) if meas else "-")
            err = round(abs(meas-gt),4) if meas else "-"
            ws_qm.cell(row,3,err)
            ok  = isinstance(err,float) and err<0.5
            c   = ws_qm.cell(row,4,"✅" if ok else "❌")
            c.fill = _copy.copy(PASS_FILL if ok else FAIL_FILL)

    # ── 평활도 요약 소시트 ──
    flat_res = [r for r in all_results if r.get("exp_type")=="flatness"]
    if flat_res:
        base_row = len(all_results)+6+len(vert_res)+3
        ws_qm.cell(base_row,1,"[평활도 요약]").font=_copy.copy(HDR_FONT)
        ws_qm.cell(base_row+1,1,"bump GT(mm)")
        ws_qm.cell(base_row+1,2,"최대잔차(mm)")
        ws_qm.cell(base_row+1,3,"클러스터")
        ws_qm.cell(base_row+1,4,"판정")
        for k,r in enumerate(flat_res):
            gt  = r.get("experiment","").replace("bump_","").replace("mm","")
            qm  = r.get("quality_metrics",{})
            row = base_row+2+k
            ws_qm.cell(row,1,gt)
            ws_qm.cell(row,2,qm.get("flatness_max_mm","-"))
            ws_qm.cell(row,3,qm.get("defect_clusters","-"))
            fp  = qm.get("flatness_pass")
            c   = ws_qm.cell(row,4,"✅평탄" if fp else "❌요철")
            if fp is not None:
                c.fill=_copy.copy(PASS_FILL if fp else FAIL_FILL)

    # ══════════════════════════════════════════
    # 시트7b: 기선 → 픽셀검출정확도 (V/H 분리)
    # ══════════════════════════════════════════
    # 목적: baseline이 "깊이 정확도"가 아니라 "픽셀 검출 자체"에 미치는
    # 영향을 직접 확인. baseline 방향(카메라가 옆으로 벌어지는 축)과
    # 평행/수직한 V선·H선이 원근 영향을 다르게 받을 수 있어 분리해서 본다.
    base_only = [r for r in all_results if r.get("exp_type") == "baseline"]
    if base_only:
        ws3b = wb.create_sheet("기선_픽셀정확도")
        ws3b.freeze_panes = "A2"
        hdrs3b = ["기선(mm)", "거리(m)", "전체pixel_rmse(px)", "전체pixel_mean(px)",
                  "V선_rmse(px)", "H선_rmse(px)", "비교점수"]
        for ci, h in enumerate(hdrs3b, 1):
            _hdr(ws3b, 1, ci, h, 16)
        base_only_sorted = sorted(base_only, key=lambda r: r.get("baseline_mm", 0))
        for ri, r in enumerate(base_only_sorted, 2):
            pa_r = r.get("pixel_accuracy", {})
            row = [r.get("baseline_mm", ""), r.get("standoff_m", ""),
                   pa_r.get("pixel_rmse_px"), pa_r.get("pixel_mean_px"),
                   pa_r.get("v_rmse_px"),
                   pa_r.get("h_rmse_px"), pa_r.get("n_compared")]
            for ci, val in enumerate(row, 1):
                c = ws3b.cell(ri, ci,
                              round(val, 4) if isinstance(val, float) else val)
                c.alignment = CENTER; c.border = BORDER
                if ri % 2 == 0: c.fill = _copy.copy(ALT_FILL)
        # 최소 pixel_rmse 기선 강조
        valid_rows = [(ri, r.get("pixel_accuracy", {}).get("pixel_rmse_px"))
                      for ri, r in enumerate(base_only_sorted, 2)
                      if r.get("pixel_accuracy", {}).get("pixel_rmse_px") is not None]
        if valid_rows:
            best_ri = min(valid_rows, key=lambda x: x[1])[0]
            for ci in range(1, len(hdrs3b) + 1):
                ws3b.cell(best_ri, ci).fill = _copy.copy(PASS_FILL)

    # ══════════════════════════════════════════
    # 시트7: 기선×거리 σZ 히트맵 — RMSE(mean) 형태로 표시
    # ══════════════════════════════════════════
    # exp_type이 grid(전체격자)/baseline(d=1.0 단면)/distance(b=150 단면)
    # 셋 중 해당 (b,d) 조합을 가진 걸 모두 모아 하나의 조회 테이블로 사용.
    # → --full-grid 없이 돌렸으면 자동으로 십자형만 채워지고,
    #   --full-grid로 돌렸으면 36칸(교차점 포함) 전체가 채워짐.
    lookup = {}
    for r in all_results:
        b = r.get("baseline_mm"); d = r.get("standoff_m")
        if b is None or d is None: continue
        lookup[(b, d)] = r

    ws3 = wb.create_sheet("σZ_히트맵")
    b_vals = sorted(set(r.get("baseline_mm",150) for r in all_results))
    d_vals = sorted(set(r.get("standoff_m",1.0) for r in all_results))

    ws3.cell(1,1,"σZ(mm) RMSE(mean) ↓거리 →기선")
    ws3.cell(1,1).fill=_copy.copy(HDR_FILL); ws3.cell(1,1).font=_copy.copy(HDR_FONT)
    ws3.cell(1,1).alignment=CENTER
    for ci,b in enumerate(b_vals,2):
        _hdr(ws3,1,ci,f"{b:.0f}mm",14)
    for ri,d in enumerate(d_vals,2):
        _hdr(ws3,ri,1,f"{d}m",8)
        for ci,b in enumerate(b_vals,2):
            r = lookup.get((b, d))
            da_r = r.get("depth_accuracy", {}) if r else {}
            rmse_v = da_r.get("sigma_z_mm")
            mean_v = da_r.get("mean_z_err_mm")
            if rmse_v is None:
                disp = "N/A"
            elif mean_v is None:
                disp = round(rmse_v, 3)
            else:
                disp = f"{rmse_v:.2f}({mean_v:.2f})"
            ok = rmse_v is not None and rmse_v <= 2.0
            c=ws3.cell(ri,ci,disp)
            c.alignment=CENTER; c.border=BORDER
            c.fill = _copy.copy(PASS_FILL) if ok else (_copy.copy(FAIL_FILL) if rmse_v is not None else PatternFill())
    ws3.cell(len(d_vals)+3, 1,
             "※ 셀 표기: RMSE(mean) — 판정/비교 기준은 RMSE(앞 숫자). "
             "N/A는 해당 조합이 미실행(--full-grid로 채울 수 있음)")

    # ══════════════════════════════════════════
    # 시트8: 카메라/레이저 위치좌표
    # ══════════════════════════════════════════
    ws4 = wb.create_sheet("위치좌표")
    ws4.freeze_panes = "A2"

    hdrs4 = ["실험라벨", "exp_type", "기선(mm)", "거리(m)",
             "벽중심_X", "벽중심_Y", "벽중심_Z",
             "레이저_X", "레이저_Y", "레이저_Z",
             "카메라_X", "카메라_Y", "카메라_Z",
             "forward_X", "forward_Y", "forward_Z"]
    widths4 = [18,14,10,8, 10,10,10, 10,10,10, 10,10,10, 10,10,10]
    for ci,(h,w) in enumerate(zip(hdrs4,widths4),1):
        _hdr(ws4,1,ci,h,w)

    for ri, res in enumerate(all_results, 2):
        cw = res.get("camera_world", {})
        wc = cw.get("wall_center_m", [None,None,None])
        lp = cw.get("laser_position_m", [None,None,None])
        cp = cw.get("position_m", [None,None,None])
        fd = cw.get("forward_dir", [None,None,None])

        vals = [
            res.get("experiment",""),
            res.get("exp_type",""),
            res.get("baseline_mm",""),
            res.get("standoff_m",""),
            *[round(v,4) if isinstance(v,(int,float)) else v for v in wc],
            *[round(v,4) if isinstance(v,(int,float)) else v for v in lp],
            *[round(v,4) if isinstance(v,(int,float)) else v for v in cp],
            *[round(v,4) if isinstance(v,(int,float)) else v for v in fd],
        ]
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(vals, 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            c.alignment = CENTER; c.border = BORDER
            if fill: c.fill = _copy.copy(fill)

    # ══════════════════════════════════════════
    # 시트9: 실험별 이미지 (overlay.png 썸네일)
    # 이미지 삽입이 실패해도 나머지 시트는 정상 저장되도록 전체를 보호
    # ══════════════════════════════════════════
    try:
        ws5 = wb.create_sheet("이미지")
        ws5.freeze_panes = "A2"
        hdrs5 = ["실험라벨", "exp_type", "기선(mm)", "거리(m)",
                 "σZ(mm)", "overlay (레이저+raycast)"]
        widths5 = [18,14,10,8,10,42]
        for ci,(h,w) in enumerate(zip(hdrs5,widths5),1):
            _hdr(ws5,1,ci,h,w)

        THUMB_W_PX = 280   # 엑셀 셀에 들어갈 썸네일 가로 px
        ROW_HEIGHT = 130   # 행 높이(포인트)

        for ri, res in enumerate(all_results, 2):
            label = res.get("experiment","")
            da = res.get("depth_accuracy", {})
            ws5.row_dimensions[ri].height = ROW_HEIGHT

            vals = [label, res.get("exp_type",""),
                    res.get("baseline_mm",""), res.get("standoff_m",""),
                    round(da.get("sigma_z_mm"),3) if da.get("sigma_z_mm") else None]
            for ci, val in enumerate(vals, 1):
                c = ws5.cell(row=ri, column=ci, value=val)
                c.alignment = CENTER; c.border = BORDER

            img_dir = os.path.join(out_dir, label)
            fpath = os.path.join(img_dir, "overlay.png")
            if os.path.exists(fpath):
                try:
                    pil_img = Image.open(fpath).convert("RGB")
                    ratio = THUMB_W_PX / pil_img.width
                    thumb = pil_img.resize(
                        (THUMB_W_PX, int(pil_img.height*ratio)))
                    thumb_path = os.path.join(img_dir, "_thumb_overlay.png")
                    thumb.save(thumb_path, format="PNG")
                    pil_img.close(); thumb.close()
                    xlimg = XLImage(thumb_path)
                    ws5.add_image(xlimg, f"F{ri}")
                except Exception as e:
                    ws5.cell(row=ri, column=6, value=f"[로드실패:{e}]")
            else:
                ws5.cell(row=ri, column=6, value="(없음)")
    except Exception as e:
        LOG(f"  [경고] 이미지 시트 생성 실패, 나머지 시트만 저장: {e}")

    xp = os.path.join(out_dir, "실험결과.xlsx")
    try:
        wb.save(xp)
        LOG(f"엑셀 저장 → {xp}")
    except Exception as e:
        LOG(f"  [경고] 엑셀 저장 실패: {e}")
        # 이미지 시트 제거 후 재시도 (용량/메모리 문제 대응)
        try:
            if "이미지" in wb.sheetnames:
                del wb["이미지"]
            xp2 = os.path.join(out_dir, "실험결과_이미지없음.xlsx")
            wb.save(xp2)
            LOG(f"  → 이미지 없이 재저장 성공: {xp2}")
        except Exception as e2:
            LOG(f"  [오류] 재저장도 실패: {e2}")



# =====================================================================
# 메인
# =====================================================================
def main():
    parser = argparse.ArgumentParser(
        description="기선×거리×조건 실험")
    parser.add_argument("--algo-detect",      default=None,
                        help="[A] 선검출 알고리즘 .py 경로 (detect() 함수)")
    parser.add_argument("--algo-triangulate", default=None,
                        help="[B] 삼각측량 알고리즘 .py 경로 (triangulate() 함수)")
    parser.add_argument("--full-grid", action="store_true",
                        help="기선×거리 전체 격자(36조합 중 교차점 제외 25개) 추가 실행. "
                             "기본은 십자형 단면(실험1+5)만 실행하여 시간 절약.")
    parser.add_argument("--diff-image", action="store_true",
                        help="[방안4] 차영상 모드. 레이저 ON/OFF 프레임 차분으로 "
                             "배경광 제거 (현장 직사광 대비). A_선검출이 "
                             "laser_off_image를 지원해야 함.")
    args = parser.parse_args()

    fn_det = _default_detect
    fn_tri = _default_triangulate
    if args.algo_detect:
        spec = importlib.util.spec_from_file_location("_A", args.algo_detect)
        m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m)
        fn_det = m.detect
        LOG(f"[A] 외부 알고리즘: {args.algo_detect}")
    if args.algo_triangulate:
        spec = importlib.util.spec_from_file_location("_B", args.algo_triangulate)
        m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m)
        fn_tri = m.triangulate
        LOG(f"[B] 외부 알고리즘: {args.algo_triangulate}")

    stage, world    = _setup_scene()
    try: world.reset()
    except Exception: pass
    camera          = _setup_camera(stage)
    line_angles     = _make_line_angles(
        GRID_PARAMS["n_vertical"], GRID_PARAMS["n_horizontal"],
        GRID_PARAMS["fov_deg"])
    _wait(world, 30)
    # GT 로드
    import json as _json
    _GT_JSON = "/home/develop/Desktop/laser_grid_test_4/inspection_ground_truth_realistic.json"
    import os as _os
    gt_full = _json.load(open(_GT_JSON,encoding="utf-8")) if _os.path.exists(_GT_JSON) else {}
    run_sweep(stage, world, camera, line_angles, fn_det, fn_tri, gt_full,
             full_grid=args.full_grid, use_diff_image=args.diff_image)


if __name__ == "__main__":
    main()
